"""
🔐 Процессор массового сброса паролей v2.1
Автоматическая отправка запросов на восстановление пароля
Поддержка: Max.Credit, Свои Люди

Changelog v2.1:
- Добавлено принудительное обновление страницы после каждой отправки
- Исправлена проблема stale element reference
- Улучшена стабильность работы
"""

import time
import traceback
from io import BytesIO
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# Telegram уведомления
try:
    from telegram_manager import send_notification_sync, is_bot_available
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    def send_notification_sync(msg): pass
    def is_bot_available(): return False

try:
    import easyocr
    from PIL import Image, ImageEnhance
    import numpy as np
    CAPTCHA_AVAILABLE = True
except ImportError:
    CAPTCHA_AVAILABLE = False
    print("⚠️ Для распознавания капчи нужны: easyocr, PIL, numpy")


# =============================================================================
# КОНФИГУРАЦИИ САЙТОВ
# =============================================================================

SITE_CONFIGS = {
    "max.credit": {
        "name": "Max.Credit",
        "restore_url": "https://www.max.credit/auth/restore",
        "xpath_phone": "/html/body/div[2]/div/form/div[1]/input",
        "xpath_captcha_img": "/html/body/div[2]/div/form/div[2]/img",
        "xpath_captcha_input": "/html/body/div[2]/div/form/div[2]/input",
        "xpath_submit_with_captcha": "/html/body/div[2]/div/form/div[3]/button",
        "xpath_submit_no_captcha": "/html/body/div[2]/div/form/div[2]/button"
    },
    "svoi-ludi.ru": {
        "name": "Свои Люди",
        "restore_url": "https://svoi-ludi.ru/auth/restore",
        "xpath_phone": "/html/body/div[2]/div[3]/div/div/form/div[1]/input",
        "xpath_captcha_img": "/html/body/div[2]/div[3]/div/div/form/div[2]/img",
        "xpath_captcha_input": "/html/body/div[2]/div[3]/div/div/form/div[2]/input",
        "xpath_submit_with_captcha": "/html/body/div[2]/div[3]/div/div/form/div[3]/button",
        "xpath_submit_no_captcha": "/html/body/div[2]/div[3]/div/div/form/div[2]/button"
    }
}


def detect_site(user_link):
    """
    Определяет сайт по ссылке пользователя
    
    Args:
        user_link: Ссылка из настроек пользователя
    
    Returns:
        dict: Конфигурация сайта
    """
    if not user_link:
        # По умолчанию Max.Credit
        return SITE_CONFIGS["max.credit"]
    
    user_link_lower = user_link.lower()
    
    if "svoi-ludi" in user_link_lower or "свои-люди" in user_link_lower:
        return SITE_CONFIGS["svoi-ludi.ru"]
    elif "max.credit" in user_link_lower or "макс.кредит" in user_link_lower:
        return SITE_CONFIGS["max.credit"]
    else:
        # По умолчанию Max.Credit
        return SITE_CONFIGS["max.credit"]


# =============================================================================
# РАСПОЗНАВАНИЕ КАПЧИ
# =============================================================================

# Глобальный reader для EasyOCR
_easyocr_reader = None


def get_easyocr_reader():
    """Ленивая инициализация EasyOCR reader"""
    global _easyocr_reader
    if _easyocr_reader is None:
        if not CAPTCHA_AVAILABLE:
            raise ImportError("EasyOCR и зависимости недоступны!")
        _easyocr_reader = easyocr.Reader(['en'], gpu=False)
    return _easyocr_reader


def solve_captcha(driver, captcha_element, logger_func):
    """
    Распознаёт капчу (универсально для обоих сайтов)
    
    Args:
        driver: WebDriver
        captcha_element: Элемент с картинкой капчи
        logger_func: Функция логирования
    
    Returns:
        str: Распознанная капча или None
    """
    try:
        if not CAPTCHA_AVAILABLE:
            logger_func("❌ Библиотеки для распознавания капчи недоступны!")
            return None
        
        # Скриншот капчи
        captcha_png = captcha_element.screenshot_as_png
        img = Image.open(BytesIO(captcha_png)).convert('RGB')
        
        # Увеличиваем в 3 раза
        img = img.resize((img.width * 3, img.height * 3), Image.LANCZOS)
        
        # Контраст + резкость
        img = ImageEnhance.Contrast(img).enhance(3.0)
        img = ImageEnhance.Sharpness(img).enhance(2.0)
        
        # Распознавание
        reader = get_easyocr_reader()
        result = reader.readtext(
            np.array(img),
            allowlist='0123456789',
            detail=0,
            paragraph=False
        )
        
        if result:
            captcha_text = ''.join(result)
            captcha_text = ''.join(filter(str.isdigit, captcha_text))
            
            if len(captcha_text) == 6:
                logger_func(f"✅ Капча распознана: {captcha_text}")
                return captcha_text
            else:
                logger_func(f"⚠️ Распознано {len(captcha_text)} цифр вместо 6: {captcha_text}")
                return None
        else:
            logger_func("❌ Капча не распознана")
            return None
            
    except Exception as e:
        logger_func(f"❌ Ошибка распознавания капчи: {e}")
        return None


# =============================================================================
# ОСНОВНОЙ ПРОЦЕССОР
# =============================================================================

def process_password_reset(excel_file_path, logger_func, stop_flag=None, user_link=None):
    """
    Массовая отправка запросов на восстановление пароля
    
    Args:
        excel_file_path: Путь к Excel файлу с номерами телефонов
        logger_func: Функция логирования
        stop_flag: threading.Event для остановки процесса
        user_link: Ссылка из настроек пользователя (для определения сайта)
    """
    
    # Определяем сайт
    site_config = detect_site(user_link)
    site_name = site_config["name"]
    
    logger_func(f"🔐 Запуск процесса восстановления паролей...")
    logger_func(f"🌐 Сайт: {site_name}")
    logger_func(f"🔗 URL восстановления: {site_config['restore_url']}")

    driver = None
    total_sent = 0
    total_errors = 0

    # Получаем XPath из конфигурации
    RESTORE_URL = site_config["restore_url"]
    XPATH_PHONE_INPUT = site_config["xpath_phone"]
    XPATH_CAPTCHA_IMAGE = site_config["xpath_captcha_img"]
    XPATH_CAPTCHA_INPUT = site_config["xpath_captcha_input"]
    XPATH_SUBMIT_WITH_CAPTCHA = site_config["xpath_submit_with_captcha"]
    XPATH_SUBMIT_NO_CAPTCHA = site_config["xpath_submit_no_captcha"]

    RESTART_INTERVAL = 30  # каждые 30 номеров обновляем браузер
    DELAY_BETWEEN_REQUESTS = 60

    def start_browser():
        """Запуск невидимого браузера с защитой от race condition"""
        max_attempts = 3
        
        for attempt in range(max_attempts):
            try:
                options = webdriver.ChromeOptions()
                options.add_argument('--headless')
                options.add_argument('--disable-gpu')
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                options.add_argument('--window-size=1920,1080')

                service = Service(ChromeDriverManager().install())
                browser = webdriver.Chrome(service=service, options=options)
                return browser
                
            except Exception as e:
                if "version" in str(e).lower() and attempt < max_attempts - 1:
                    logger_func(f"⚠️ Попытка {attempt+1}/{max_attempts}: конфликт версий, retry через 5 сек...")
                    time.sleep(5)
                    
                    # Очистка кэша при ошибке версий
                    try:
                        import shutil
                        from pathlib import Path
                        wdm_cache = Path.home() / '.wdm'
                        if wdm_cache.exists():
                            shutil.rmtree(wdm_cache)
                            logger_func("🧹 Кэш драйвера очищен")
                    except:
                        pass
                else:
                    raise

    try:
        # Читаем Excel файл
        logger_func(f"📄 Чтение файла: {excel_file_path}")
        workbook = load_workbook(excel_file_path)
        sheet = workbook.active

        # Ищем столбец с телефонами
        phone_column = None
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value and 'телефон' in str(cell.value).lower():
                phone_column = col_idx
                logger_func(f"📌 Найден столбец: '{cell.value}' (#{col_idx})")
                break

        if not phone_column:
            logger_func("❌ Столбец 'Телефон' не найден!")
            return

        # Извлекаем телефоны
        phone_numbers = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            phone = row[phone_column - 1]
            if phone:
                phone_str = str(phone).strip()
                # Убираем всё кроме цифр
                phone_clean = ''.join(filter(str.isdigit, phone_str))
                if phone_clean:
                    phone_numbers.append(phone_clean)

        if not phone_numbers:
            logger_func(
                "❌ В файле не найдено ни одного телефона!\n"
                f"Проверьте данные в Excel"
            )
            return

        logger_func(f"📊 Найдено телефонов: {len(phone_numbers)}")
        
        # 📱 Уведомление о запуске
        if TELEGRAM_AVAILABLE and is_bot_available():
            send_notification_sync(
                f"🔐 <b>СБРОС ПАРОЛЕЙ ЗАПУЩЕН</b>\n\n"
                f"🌐 Сайт: <b>{site_name}</b>\n"
                f"📱 Номеров в очереди: <b>{len(phone_numbers)}</b>\n"
                f"🤖 Режим: Headless (невидимый браузер)\n"
                f"🔄 Автоперезапуск: каждые 30 номеров"
            )

        # Запуск браузера
        logger_func("🌐 Запуск headless браузера...")
        driver = start_browser()
        wait = WebDriverWait(driver, 20)

        driver.get(RESTORE_URL)
        time.sleep(3)

        # Основной цикл
        for idx, phone in enumerate(phone_numbers, 1):

            # Остановка фоновой задачи
            if stop_flag and stop_flag.is_set():
                logger_func("⛔ Остановка по команде пользователя")
                
                # 📱 Уведомление об остановке
                if TELEGRAM_AVAILABLE and is_bot_available():
                    send_notification_sync(
                        f"⏹️ <b>СБРОС ПАРОЛЕЙ ОСТАНОВЛЕН</b>\n\n"
                        f"Остановлено вручную пользователем\n"
                        f"📊 Обработано: {idx - 1}/{len(phone_numbers)}"
                    )
                break

            # Перезапуск браузера каждые N номеров
            if idx > 1 and (idx - 1) % RESTART_INTERVAL == 0:
                logger_func(f"♻ Перезапуск браузера (обработано {idx - 1})...")
                try:
                    driver.quit()
                except:
                    pass
                time.sleep(3)
                driver = start_browser()
                wait = WebDriverWait(driver, 20)
                driver.get(RESTORE_URL)
                time.sleep(3)
                logger_func("🔄 Браузер перезапущен")

            logger_func(f"\n{'='*60}")
            logger_func(f"📨 Отправка {idx}/{len(phone_numbers)}: {phone}")

            max_retries = 3
            retry = 0

            while retry < max_retries:

                if stop_flag and stop_flag.is_set():
                    logger_func("⛔ Остановка внутри цикла")
                    break

                try:
                    # Ввод телефона
                    phone_input = wait.until(EC.presence_of_element_located((By.XPATH, XPATH_PHONE_INPUT)))
                    phone_input.clear()
                    phone_input.send_keys(phone)
                    time.sleep(0.5)

                    logger_func(f"✓ Телефон введён: {phone}")

                    # Проверяем наличие капчи
                    try:
                        captcha_img = driver.find_element(By.XPATH, XPATH_CAPTCHA_IMAGE)
                        has_captcha = True
                        logger_func("🔍 Капча обнаружена")
                    except:
                        has_captcha = False
                        logger_func("👌 Капчи нет")

                    if has_captcha:
                        logger_func("🤖 Распознаём капчу...")
                        captcha_code = solve_captcha(driver, captcha_img, logger_func)

                        if not captcha_code or len(captcha_code) != 6:
                            retry += 1
                            logger_func(f"⚠ Неверная капча ({retry}/{max_retries})")
                            if retry < max_retries:
                                driver.refresh()
                                time.sleep(2)
                                continue
                            else:
                                logger_func("❌ Пропуск — капча не распознана")
                                total_errors += 1
                                break

                        captcha_input = wait.until(EC.presence_of_element_located((By.XPATH, XPATH_CAPTCHA_INPUT)))
                        captcha_input.clear()
                        captcha_input.send_keys(captcha_code)
                        time.sleep(0.5)

                        logger_func(f"🔑 Капча введена: {captcha_code}")
                        submit_xpath = XPATH_SUBMIT_WITH_CAPTCHA
                    else:
                        submit_xpath = XPATH_SUBMIT_NO_CAPTCHA

                    # ⚡ СТАБИЛИЗАЦИЯ: Даём странице время на полную перерисовку после ввода
                    time.sleep(1.5)
                    
                    # Отправляем запрос с защитой от stale element
                    logger_func("▶ Поиск кнопки отправки...")
                    
                    # Повторяем поиск кнопки до 3 раз с коротким интервалом (защита от stale element)
                    submit_btn = None
                    for attempt in range(3):
                        try:
                            submit_btn = wait.until(EC.element_to_be_clickable((By.XPATH, submit_xpath)))
                            break  # Элемент найден и кликабелен
                        except Exception as btn_error:
                            if attempt < 2:  # Ещё есть попытки
                                time.sleep(1)  # Короткая пауза перед повтором
                            else:
                                raise btn_error  # Исчерпаны попытки - выбрасываем ошибку
                    
                    if submit_btn:
                        submit_btn.click()
                        time.sleep(2)

                    # Проверка успеха
                    try:
                        # Проверяем URL или сообщение об успехе
                        current_url = driver.current_url
                        if "success" in current_url.lower() or driver.current_url != RESTORE_URL:
                            logger_func("✅ Запрос отправлен успешно")
                            total_sent += 1
                            break
                        else:
                            # Ищем сообщение об ошибке
                            try:
                                error_msg = driver.find_element(By.CLASS_NAME, "error").text
                                logger_func(f"⚠️ Ошибка на странице: {error_msg}")
                            except:
                                logger_func("✅ Запрос вероятно отправлен (проверьте вручную)")
                                total_sent += 1
                            break
                    except:
                        logger_func("✅ Запрос вероятно отправлен")
                        total_sent += 1
                        break

                except Exception as e:
                    retry += 1
                    logger_func(f"❌ Ошибка ({retry}/{max_retries}): {str(e)[:100]}")
                    
                    if retry < max_retries:
                        logger_func("🔄 Обновление страницы...")
                        driver.get(RESTORE_URL)
                        time.sleep(2)
                    else:
                        logger_func("❌ Пропуск номера после исчерпания попыток")
                        total_errors += 1
                        break

            # Пауза между запросами + обновление страницы для избежания stale elements
            if idx < len(phone_numbers):
                logger_func(f"⏳ Ожидание {DELAY_BETWEEN_REQUESTS} секунд...")
                time.sleep(DELAY_BETWEEN_REQUESTS)
                
                # ⚡ ФИКС: Явное обновление страницы перед следующей итерацией
                # Это предотвращает ошибки stale element reference после успешной отправки
                logger_func("🔄 Обновление формы для следующей отправки...")
                driver.get(RESTORE_URL)
                time.sleep(2)

        # Финальная статистика
        logger_func(f"\n{'='*60}")
        logger_func(f"✅ ЗАВЕРШЕНО")
        logger_func(f"📊 Всего обработано: {len(phone_numbers)}")
        logger_func(f"✅ Отправлено: {total_sent}")
        logger_func(f"❌ Ошибок: {total_errors}")
        logger_func(f"{'='*60}")

        # 📱 Финальное уведомление
        if TELEGRAM_AVAILABLE and is_bot_available():
            send_notification_sync(
                f"✅ <b>СБРОС ПАРОЛЕЙ ЗАВЕРШЁН</b>\n\n"
                f"🌐 Сайт: <b>{site_name}</b>\n"
                f"📊 Всего: {len(phone_numbers)}\n"
                f"✅ Отправлено: <b>{total_sent}</b>\n"
                f"❌ Ошибок: {total_errors}\n"
                f"⏱ Завершено: {datetime.now().strftime('%H:%M:%S')}"
            )

    except Exception as e:
        error_msg = f"💥 Критическая ошибка: {str(e)}"
        logger_func(error_msg)
        logger_func(traceback.format_exc())
        
        # 📱 Критическое уведомление
        if TELEGRAM_AVAILABLE and is_bot_available():
            send_notification_sync(
                f"🚨 <b>СБРОС ПАРОЛЕЙ УПАЛ</b>\n\n"
                f"❌ {str(e)[:200]}\n\n"
                f"📊 Обработано до ошибки: {total_sent}/{len(phone_numbers) if 'phone_numbers' in locals() else '?'}"
            )

    finally:
        if driver:
            try:
                driver.quit()
                logger_func("🔴 Браузер закрыт")
            except:
                pass