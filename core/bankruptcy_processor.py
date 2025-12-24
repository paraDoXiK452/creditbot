# -*- coding: utf-8 -*-
"""
💼 Процессор банкротства
Проверка ФИО через kad.arbitr.ru с анализом регионов
"""

import time
import random
import traceback
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
from process_manager import register_driver

# Telegram уведомления
try:
    from telegram_manager import send_notification_sync, is_bot_available
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    def send_notification_sync(msg): pass
    def is_bot_available(): return False


# =============================================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =============================================================================

def extract_region_from_address(address):
    """
    Извлекает регион из полного адреса или названия суда.
    Примеры:
    "302000 Орловская Область Орёл..." -> "Орловская"
    "АС Орловской области" -> "Орловская"
    """
    if not address:
        return ""
    
    address_lower = address.lower().strip()
    
    # Убираем префиксы
    address_lower = address_lower.replace("ас ", "").replace("арбитражный суд ", "")
    
    # Список КОРНЕЙ регионов (без окончаний)
    regions = [
        "московск", "ленинградск", "свердловск", "новосибирск", 
        "ростовск", "нижегородск", "самарск", "омск", "челябинск",
        "волгоградск", "воронежск", "саратовск", "красноярск",
        "пермск", "тюменск", "иркутск", "томск", "кемеровск",
        "архангельск", "астраханск", "белгородск", "брянск",
        "владимирск", "вологодск", "ивановск",
        "калининградск", "калужск", "кировск", "костромск",
        "курганск", "курск", "липецк", "магаданск", "мурманск",
        "новгородск", "оренбургск", "орловск", "пензенск", "псковск",
        "рязанск", "смоленск", "тамбовск", "тверск", "тульск",
        "ульяновск", "ярославск", "амурск", "сахалинск",
        
        "татарстан", "башкортостан", "дагестан", "бурят", "якут",
        "чуваш", "мордов", "удмурт", "марий", "коми", "карел",
        "калмык", "тыва", "хакас", "алтай", "адыг", "кабардин",
        "карачаев", "осет", "ингуш", "чечен",
        
        "приморск", "забайкальск", "камчатск", "ставропольск", "хабаровск"
    ]
    
    # Ищем совпадение по КОРНЮ
    for region in regions:
        if region in address_lower:
            # Возвращаем с правильным окончанием
            if region == "орловск":
                return "Орловская"
            elif region == "московск":
                return "Московская"
            elif region.endswith("ск"):
                return region.capitalize() + "ая"
            else:
                return region.capitalize()
    
    return ""


def is_date_within_months(date_str, months=6):
    """Проверяет, что дата не старше N месяцев от текущей"""
    try:
        # Парсим дату формата "20.03.2025" или "03.08.2023"
        date_obj = datetime.strptime(date_str.strip(), "%d.%m.%Y")
        current_date = datetime.now()
        
        # Вычисляем разницу
        diff_months = (current_date.year - date_obj.year) * 12 + (current_date.month - date_obj.month)
        
        return diff_months <= months
    except Exception:
        # Если не удалось распарсить - считаем что подходит
        return True


def _close_any_modal(driver, logger_func, timeout=5):
    """
    Закрывает любые всплывающие окна на kad.arbitr.ru:
      1) promo-попапы (b-promo-notice, b-promo_notification_popup)
      2) системные модалки (.modal с крестиком в заголовке)
    """
    try:
        # Быстрый выход, если ничего нет
        if not driver.find_elements(By.XPATH, "//div[contains(@class,'b-promo')]") \
           and not driver.find_elements(By.XPATH, "//div[contains(@class,'modal')]"):
            return

        # === 1) НОВЫЙ ТИП ПОПАПА: b-promo_notification_popup_wrapper ===
        try:
            new_popup_xpaths = [
                "//div[contains(@class,'b-promo_notification')]//button[contains(@class,'close')]",
                "//div[contains(@class,'b-promo_notification')]//a[contains(@class,'close')]",
                "//div[contains(@class,'b-promo_notification')]//button",
                "//div[contains(@class,'b-promo_notification')]//a",
                "//div[contains(@class,'b-promo_notification')]//*[@class='close']"
            ]
            
            for xpath in new_popup_xpaths:
                try:
                    btn = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, xpath)))
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.5)
                    logger_func("Закрыт новый promo-попап (b-promo_notification)")
                    return
                except:
                    continue
        except:
            pass

        # === 2) СТАРЫЙ ТИП: b-promo-notice ===
        try:
            xp_abs = "/html/body/div[13]/div[2]/div/div/div/div/b/b/a[1]"
            btn = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, xp_abs)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center',inline:'center'});", btn)
            try:
                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, xp_abs)))
                btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
            WebDriverWait(driver, timeout).until_not(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'b-promo-notice')]"))
            )
            logger_func("Закрыт promo-попап (абсолютный XPath).")
            return
        except Exception:
            try:
                btn = WebDriverWait(driver, 2).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[@class='b-promo-notice__close']//a"))
                )
                driver.execute_script("arguments[0].click();", btn)
                WebDriverWait(driver, timeout).until_not(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'b-promo-notice')]"))
                )
                logger_func("Закрыт promo-попап (универсальный селектор).")
                return
            except Exception:
                pass

        # === 3) СИСТЕМНЫЕ МОДАЛКИ (.modal) ===
        try:
            btn = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//div[contains(@class,'modal')]//button[contains(@class,'close')]"
                    " | //div[contains(@class,'modal')]//a[contains(@class,'close')]"
                    " | //div[contains(@class,'modal')]//a[@aria-label='Close']"
                ))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center',inline:'center'});", btn)
            try:
                btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", btn)
            WebDriverWait(driver, timeout).until_not(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class,'modal')]"))
            )
            logger_func("Закрыта системная модалка.")
            return
        except Exception:
            pass
    except Exception as e:
        logger_func(f"Ошибка при закрытии окна: {e}")


def _normalize_viewport(driver):
    """Лёгкая нормализация, чтобы прибить горизонтальный скролл/дёргание ширины."""
    driver.execute_script("""
        try {
            document.documentElement.style.zoom = '1';
            document.body.style.zoom = '1';
            document.documentElement.style.removeProperty('width');
            document.body.style.removeProperty('width');
            document.documentElement.style.overflowX = 'hidden';
            document.body.style.overflowX = 'hidden';
            window.dispatchEvent(new Event('resize'));
        } catch(e) {}
    """)


def _save_debug_screenshot(driver, fio_name, logger_func, reason=""):
    """
    Сохраняет скриншот для отладки с понятным именем.
    
    Args:
        driver: Selenium WebDriver
        fio_name: Имя клиента для имени файла
        logger_func: Функция логирования
        reason: Причина создания скриншота
    
    Returns:
        str: Путь к сохраненному скриншоту или None
    """
    try:
        # Создаем понятное имя файла
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_fio = "".join(c for c in fio_name if c.isalnum() or c in (' ', '_')).strip()
        safe_fio = safe_fio.replace(' ', '_')[:30]  # Ограничиваем длину
        
        reason_text = f"_{reason}" if reason else ""
        screenshot_path = f"debug_kad_{safe_fio}_{timestamp}{reason_text}.png"
        
        driver.save_screenshot(screenshot_path)
        logger_func(f"  📸 Скриншот сохранён: {screenshot_path}")
        return screenshot_path
    except Exception as e:
        logger_func(f"  ⚠️ Не удалось сохранить скриншот: {e}")
        return None


def _wait_for_results(driver, logger_func, fio_name="unknown", timeout=60):
    """
    Ждёт результатов поиска при динамической загрузке.
    Возвращает: "none" | "found" | "unknown" | "captcha"
    """
    logger_func("  Ожидание динамической загрузки результатов...")
    
    # Имитируем действия пользователя - скроллим немного
    try:
        driver.execute_script("window.scrollTo(0, 100);")
        time.sleep(0.5)
        driver.execute_script("window.scrollTo(0, 0);")
    except:
        pass
    
    time.sleep(3)
    
    start_time = time.time()
    last_html = ""
    stable_count = 0
    
    # Ждем стабилизации DOM
    while time.time() - start_time < timeout:
        try:
            current_html = driver.find_element(By.TAG_NAME, "body").get_attribute("innerHTML")
            
            if current_html == last_html:
                stable_count += 1
                if stable_count >= 2:
                    logger_func("  DOM стабилизировался, проверяем результаты...")
                    break
            else:
                stable_count = 0
                last_html = current_html
            
            time.sleep(1)
        except Exception as e:
            logger_func(f"  Ошибка проверки DOM: {e}")
            break
    
    if time.time() - start_time >= timeout:
        logger_func(f"  ⏱ Таймаут ({timeout}с)")
        _save_debug_screenshot(driver, fio_name, logger_func, "timeout")
        return "unknown"
    
    time.sleep(2)
    
    # === ПРОВЕРКА НА КАПЧУ ===
    # СТРОГИЙ поиск конкретных фраз капчи, без ложных срабатываний
    try:
        # Ищем конкретные фразы капчи
        captcha_xpaths = [
            "//*[contains(text(), 'Введите символы') or contains(text(), 'введите символы')]",
            "//*[contains(text(), 'Подтвердите, что вы не робот') or contains(text(), 'не робот')]",
            "//*[contains(text(), 'reCAPTCHA') or contains(text(), 'recaptcha')]",
            "//*[contains(text(), 'капча')]",
            "//img[contains(@src, 'captcha')]",
            "//div[contains(@class, 'captcha')]"
        ]
        
        for xpath in captcha_xpaths:
            elements = driver.find_elements(By.XPATH, xpath)
            if elements and any(elem.is_displayed() for elem in elements):
                logger_func("  🚨 ОБНАРУЖЕНА КАПЧА! Сайт заблокировал бота")
                _save_debug_screenshot(driver, fio_name, logger_func, "CAPTCHA")
                return "captcha"
    except Exception as e:
        logger_func(f"  Ошибка проверки капчи: {e}")
    
    # === ПРОВЕРКА РЕЗУЛЬТАТОВ ===
    
    # 1) Проверяем блок "Нет результатов" - ВИДИМЫЙ ЛИ ОН
    try:
        no_results_block = driver.find_element(By.CLASS_NAME, "b-noResults")
        classes = no_results_block.get_attribute("class")
        style = no_results_block.get_attribute("style")
        
        logger_func(f"  Блок 'Нет результатов': classes={classes}, style={style}")
        
        if "g-hidden" not in classes and "display: none" not in (style or ""):
            logger_func("  ✗ Блок 'Нет результатов' ВИДИМЫЙ")
            return "none"
        else:
            logger_func("  Блок 'Нет результатов' скрыт")
    except Exception as e:
        logger_func(f"  Блок 'Нет результатов' не найден: {e}")
    
    # 2) Проверяем блок с результатами - ВИДИМЫЙ ЛИ ОН
    try:
        results_block = driver.find_element(By.CLASS_NAME, "b-results")
        classes = results_block.get_attribute("class")
        style = results_block.get_attribute("style")
        
        logger_func(f"  Блок результатов: classes={classes}, style={style}")
        
        if "g-hidden" not in classes and "display: none" not in (style or ""):
            logger_func("  ✓ Блок результатов ВИДИМЫЙ")
            return "found"
        else:
            logger_func("  Блок результатов скрыт")
    except Exception as e:
        logger_func(f"  Блок результатов не найден: {e}")
    
    # 3) Дополнительная проверка по тексту "Найдено X дел"
    try:
        found_text_elements = driver.find_elements(By.XPATH, "//*[contains(text(), 'Найдено') and contains(text(), 'дел')]")
        for elem in found_text_elements:
            if elem.is_displayed():
                text = elem.text
                logger_func(f"  ✓ Найден видимый текст: '{text}'")
                return "found"
    except Exception as e:
        logger_func(f"  Проверка текста 'Найдено': {e}")
    
    # 4) Проверка таблицы результатов напрямую
    try:
        tables = driver.find_elements(By.XPATH, "//table[contains(@class, 'b-cases')]")
        for table in tables:
            if table.is_displayed():
                rows = table.find_elements(By.TAG_NAME, "tr")
                if len(rows) > 1:  # Больше чем просто заголовок
                    logger_func(f"  ✓ Найдена таблица с {len(rows)} строками")
                    return "found"
    except Exception as e:
        logger_func(f"  Проверка таблицы: {e}")
    
    # 5) Результат не определен - делаем финальный скриншот
    logger_func("  ⚠ Результат не определен - возможно сайт заблокировал бота")
    _save_debug_screenshot(driver, fio_name, logger_func, "unknown")
    return "unknown"


# =============================================================================
# ОСНОВНАЯ ФУНКЦИЯ ПРОВЕРКИ
# =============================================================================

def check_bankruptcy_list(excel_file_path, logger_func, stop_flag=None, progress_callback=None):
    """
    Функция для проверки списка ФИО на банкротство через kad.arbitr.ru
    
    Args:
        excel_file_path: Путь к Excel файлу с ФИО
        logger_func: Функция логирования
        stop_flag: Флаг остановки (threading.Event)
        progress_callback: Колбек для обновления прогресса (current, total)
    
    Returns:
        str: Путь к сохраненному файлу с результатами или None
    """
    logger_func("Запуск процесса проверки банкротства.")
    logger_func(f"Чтение файла: {excel_file_path}")
    
    driver = None
    
    try:
        # Чтение Excel
        wb = load_workbook(excel_file_path)
        sheet = wb.active
        
        # Поиск столбцов
        fio_column = None
        address_column = None
        
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                cell_lower = str(cell.value).lower()
                if "фио" in cell_lower:
                    fio_column = col_idx
                    logger_func(f"Найден столбец с ФИО: {cell.value} (столбец {col_idx})")
                elif "адрес" in cell_lower or "прописк" in cell_lower:
                    address_column = col_idx
                    logger_func(f"Найден столбец с адресом: {cell.value} (столбец {col_idx})")
        
        if not fio_column:
            logger_func("ОШИБКА: Столбец 'ФИО' не найден!")
            
            # 📱 Критическое уведомление
            if TELEGRAM_AVAILABLE and is_bot_available():
                send_notification_sync(
                    f"🚨 <b>БАНКРОТСТВО УПАЛО</b>\n\n"
                    f"❌ Столбец 'ФИО' не найден в Excel\n\n"
                    f"Проверьте формат файла"
                )
            return None
        
        if not address_column:
            logger_func("ПРЕДУПРЕЖДЕНИЕ: Столбец с адресом не найден.")
        
        # Формируем входные данные
        clients_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            fio = row[fio_column - 1] if fio_column else None
            address = row[address_column - 1] if address_column else None
            
            if fio and str(fio).strip():
                full_address = str(address).strip() if address else ""
                region_extracted = extract_region_from_address(full_address)
                
                clients_data.append({
                    'fio': str(fio).strip(),
                    'address': full_address,
                    'region': region_extracted
                })
        
        if not clients_data:
            logger_func("ОШИБКА: Не найдено ни одного ФИО в файле!")
            
            # 📱 Критическое уведомление
            if TELEGRAM_AVAILABLE and is_bot_available():
                send_notification_sync(
                    f"🚨 <b>БАНКРОТСТВО УПАЛО</b>\n\n"
                    f"❌ Не найдено ни одного ФИО в файле\n\n"
                    f"Проверьте данные в Excel"
                )
            return None
        
        logger_func(f"Найдено {len(clients_data)} записей для проверки")
        
        # 📱 Уведомление о запуске
        if TELEGRAM_AVAILABLE and is_bot_available():
            send_notification_sync(
                f"💼 <b>БАНКРОТСТВО ЗАПУЩЕНО</b>\n\n"
                f"📝 Записей для проверки: <b>{len(clients_data)}</b>\n"
                f"🌐 Сайт: kad.arbitr.ru\n"
                f"🤖 Режим: Stealth + Headless"
            )
        
        # Инициализация браузера (selenium-stealth для обхода детекта)
        logger_func("Запуск браузера в headless stealth режиме...")
        try:
            from selenium_stealth import stealth
            
            options = webdriver.ChromeOptions()
            options.add_argument('--headless=new')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--window-size=1280,900')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            logger_func("Инициализация ChromeDriver...")
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=options)
            
            # Применяем selenium-stealth
            stealth(driver,
                languages=["ru-RU", "ru"],
                vendor="Google Inc.",
                platform="Win32",
                webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine",
                fix_hairline=True,
            )
            
            register_driver(driver)
            wait = WebDriverWait(driver, 15)
            logger_func("✅ Браузер запущен (selenium-stealth)")
            
        except ImportError as ie:
            logger_func(f"⚠️ selenium-stealth не установлен: {ie}")
            logger_func("Пробую undetected_chromedriver...")
            
            try:
                import undetected_chromedriver as uc
                
                options = uc.ChromeOptions()
                options.add_argument('--headless=new')
                options.add_argument('--disable-gpu')
                options.add_argument('--no-sandbox')
                options.add_argument('--disable-dev-shm-usage')
                options.add_argument('--window-size=1280,900')
                
                driver = uc.Chrome(options=options, headless=True)
                register_driver(driver)
                wait = WebDriverWait(driver, 15)
                logger_func("✅ Браузер запущен (undetected)")
                
            except Exception as e2:
                logger_func(f"❌ Критическая ошибка: {e2}")
                logger_func(traceback.format_exc())
                
                # 📱 Критическое уведомление
                if TELEGRAM_AVAILABLE and is_bot_available():
                    send_notification_sync(
                        f"🚨 <b>БАНКРОТСТВО УПАЛО</b>\n\n"
                        f"❌ Не удалось запустить браузер\n\n"
                        f"Установите: pip install selenium-stealth"
                    )
                raise
        
        # Открываем сайт
        driver.get("https://kad.arbitr.ru")
        logger_func("Страница загружена, имитируем действия пользователя...")
        
        # Имитация действий пользователя
        time.sleep(random.uniform(2.0, 4.0))  # Случайная задержка
        driver.execute_script("window.scrollTo(0, 300);")  # Скролл вниз
        time.sleep(random.uniform(0.5, 1.5))
        driver.execute_script("window.scrollTo(0, 0);")  # Скролл вверх
        time.sleep(1)
        
        _normalize_viewport(driver)
        _close_any_modal(driver, logger_func, timeout=8)
        
        # Локаторы на странице поиска
        XPATH_FIO_INPUT = "/html/body/div[1]/div[1]/div[1]/dl/dd/div[1]/div/textarea"
        XPATH_SEARCH_BUTTON = "/html/body/div[1]/div[1]/div[1]/dl/dd/div[7]/div[1]/div/button"
        
        # СОЗДАЁМ временные списки для накопления результатов
        results_with_case = []  # Совпадения с номером дела
        results_without_case = []  # Требует проверки
        
        # Счётчики
        total_success = 0
        total_checks_needed = 0
        total_clean = 0
        
        # Основной цикл
        for idx, client in enumerate(clients_data, start=1):
            if stop_flag and stop_flag.is_set():
                logger_func("Получена команда остановки.")
                
                # 📱 Уведомление об остановке
                if TELEGRAM_AVAILABLE and is_bot_available():
                    send_notification_sync(
                        f"⏹️ <b>БАНКРОТСТВО ОСТАНОВЛЕНО</b>\n\n"
                        f"Остановлено вручную пользователем\n"
                        f"📊 Обработано: {idx - 1}/{len(clients_data)}"
                    )
                break
            
            # Обновляем прогресс
            if progress_callback:
                progress_callback(idx, len(clients_data))
            
            fio = client['fio']
            region_excel = client['region']
            full_address = client['address']
            
            logger_func(f"\n{'='*60}")
            logger_func(f"Проверка {idx}/{len(clients_data)}: {fio}")
            if full_address:
                logger_func(f"Адрес из Excel: {full_address[:80]}...")
            if region_excel:
                logger_func(f"Извлечен регион: {region_excel}")
            else:
                logger_func("Регион не удалось извлечь из адреса")
            
            try:
                # Закрыть окна и нормализовать
                _close_any_modal(driver, logger_func, timeout=2)
                _normalize_viewport(driver)

                # Ввод ФИО
                fio_input = wait.until(EC.presence_of_element_located((By.XPATH, XPATH_FIO_INPUT)))
                fio_input.click()
                time.sleep(0.3)
                fio_input.clear()
                time.sleep(0.3)
                fio_input.send_keys(fio)
                time.sleep(0.5)
                
                # Кнопка Найти
                search_clicked = False
                try:
                    btn = driver.find_element(By.XPATH, "//button[contains(text(), 'Найти')]")
                    driver.execute_script("arguments[0].click();", btn)
                    search_clicked = True
                except Exception:
                    try:
                        btn = driver.find_element(By.XPATH, XPATH_SEARCH_BUTTON)
                        driver.execute_script("arguments[0].click();", btn)
                        search_clicked = True
                    except Exception:
                        pass
                
                if not search_clicked:
                    logger_func("  ОШИБКА: Не удалось нажать 'Найти'")
                    continue

                # Ожидание результатов поиска
                status = _wait_for_results(driver, logger_func, fio_name=fio, timeout=75)
                
                if status == "none":
                    logger_func("  Не найдено (не записываем)")
                    total_clean += 1
                    
                elif status == "found":
                    # Ждём загрузки таблицы
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, "b-cases"))
                        )
                        time.sleep(3)
                        
                        # Ищем строки таблицы
                        table_rows = []
                        xpath_variants = [
                            "//table[@id='b-cases']/tbody/tr",
                            "//table[@id='b-cases']//tr[td]",
                            "//*[@id='b-cases']//tbody//tr"
                        ]
                        
                        for xpath_var in xpath_variants:
                            table_rows = driver.find_elements(By.XPATH, xpath_var)
                            if len(table_rows) > 0:
                                break
                        
                        logger_func(f"  Найдено записей: {len(table_rows)}")
                        
                        if len(table_rows) == 0:
                            logger_func("  Таблица пуста или строки не загрузились")
                            continue
                        
                        # ОБРАБАТЫВАЕМ ПЕРВУЮ ПОДХОДЯЩУЮ СТРОКУ (не старше 6 месяцев)
                        suitable_row = None
                        suitable_case_number = None
                        suitable_court_text = None

                        for row in table_rows:
                            try:
                                # Номер дела (столбец 1)
                                case_number_elem = row.find_element(By.XPATH, ".//td[1]//a")
                                case_number = case_number_elem.text.strip()
                                
                                # Дата дела (первая строка в ячейке дела)
                                date_elem = row.find_element(By.XPATH, ".//td[1]")
                                date_text = date_elem.text.strip().split('\n')[0]
                                
                                logger_func(f"    Проверка дела {case_number}, дата: {date_text}")
                                
                                # Проверяем дату
                                if is_date_within_months(date_text, months=6):
                                    logger_func(f"    ✓ Дата подходит (не старше 6 месяцев)")
                                    # Берём регион
                                    court_elem = row.find_element(By.XPATH, ".//td[2]/div/div[2]")
                                    court_text = court_elem.text.strip()
                                    
                                    suitable_row = row
                                    suitable_case_number = case_number
                                    suitable_court_text = court_text
                                    break
                                else:
                                    logger_func(f"    ✗ Дело слишком старое, пропускаем")
                            except Exception as e_row_check:
                                logger_func(f"    Ошибка проверки строки: {e_row_check}")
                                continue

                        if not suitable_row:
                            logger_func("  Все найденные дела старше 6 месяцев - не записываем")
                            total_clean += 1
                            continue

                        # Используем найденное подходящее дело
                        case_number = suitable_case_number
                        court_text = suitable_court_text
                        
                        try:
                            logger_func(f"    Дело {case_number}")
                            logger_func(f"      Суд/Регион на сайте: {court_text}")
                            
                            # Извлекаем регион из текста суда
                            site_region = extract_region_from_address(court_text)
                            
                            if site_region:
                                logger_func(f"      Извлечен регион: {site_region}")
                            else:
                                logger_func(f"      Не удалось извлечь регион из: {court_text}")
                            
                            # Сравнение регионов
                            match_found = False
                            
                            if region_excel and site_region:
                                excel_norm = region_excel.lower().strip()
                                site_norm = site_region.lower().strip()
                                
                                logger_func(f"      Сравнение: '{excel_norm}' vs '{site_norm}'")
                                
                                if excel_norm in site_norm or site_norm in excel_norm:
                                    logger_func(f"      СОВПАДЕНИЕ ПО РЕГИОНУ!")
                                    results_with_case.append([fio, 'Совпадение по региону', case_number, 'Точно'])
                                    total_success += 1
                                    match_found = True
                                else:
                                    logger_func(f"      Регионы не совпали")
                            else:
                                logger_func(f"      Недостаточно данных для сравнения")
                            
                            # Если не совпал - добавляем в список "Требует проверки"
                            if not match_found:
                                results_without_case.append([fio, 'Требует ручной проверки', '', 'Требует проверки'])
                                total_checks_needed += 1
                                logger_func(f"  Результат: Требует ручной проверки")
                        
                        except Exception as e_row:
                            logger_func(f"    Ошибка обработки записи: {e_row}")
                            continue
                    
                    except Exception as e_table:
                        logger_func(f"  Ошибка обработки таблицы: {e_table}")
                        logger_func(traceback.format_exc())
                
                elif status == "captcha":
                    logger_func("  🚨 ОБНАРУЖЕНА КАПЧА - сайт заблокировал!")
                    logger_func("  Завершаем проверку, чтобы не забанить IP")
                    
                    # 📱 Критическое уведомление
                    if TELEGRAM_AVAILABLE and is_bot_available():
                        send_notification_sync(
                            f"🚨 <b>БАНКРОТСТВО ОСТАНОВЛЕНО</b>\n\n"
                            f"⛔ Kad.arbitr.ru показал капчу\n"
                            f"📊 Обработано: {idx}/{len(clients_data)}\n\n"
                            f"Попробуйте позже или используйте proxy"
                        )
                    break
                
                else:
                    logger_func("  ⚠️ Результат не определен (возможно блокировка)")
                    # Добавляем паузу чтобы не забанить
                    time.sleep(5)
                
                time.sleep(random.uniform(3.0, 6.0))  # Увеличил задержку для меньшей детекции
            
            except Exception as e_check:
                logger_func(f"  Ошибка при проверке '{fio}': {e_check}")
                logger_func(traceback.format_exc())
                continue
        
        # В КОНЦЕ - создаём Excel и записываем ОТСОРТИРОВАННЫЕ результаты
        output_file = None
        if results_with_case or results_without_case:
            try:
                wb_out = Workbook()
                sheet_out = wb_out.active
                sheet_out.title = "Результаты"
                sheet_out.append(["ФИО", "Тип совпадения", "Номер дела", "Статус"])
                
                # Сначала все с номерами дел
                for result in results_with_case:
                    sheet_out.append(result)
                
                # Потом все без номеров дел
                for result in results_without_case:
                    sheet_out.append(result)
                
                output_file = f"результаты_банкротство_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                wb_out.save(output_file)
                
                logger_func(f"\n{'='*60}")
                logger_func(f"ФАЙЛ СОХРАНЁН: {output_file}")
                logger_func(f"{'='*60}")
                logger_func(f"ИТОГИ ПРОВЕРКИ БАНКРОТСТВ:")
                logger_func(f"  Всего проверено: {len(clients_data)}")
                logger_func(f"  Совпадение по региону: {total_success}")
                logger_func(f"  Требует проверки: {total_checks_needed}")
                logger_func(f"  Чистых/старых дел: {total_clean}")
                logger_func(f"{'='*60}")
                
                # 📱 Уведомление об успешном завершении
                if TELEGRAM_AVAILABLE and is_bot_available():
                    send_notification_sync(
                        f"✅ <b>БАНКРОТСТВО ЗАВЕРШЕНО</b>\n\n"
                        f"📊 Статистика:\n"
                        f"📝 Всего проверено: {len(clients_data)}\n"
                        f"✅ Совпадение по региону: <b>{total_success}</b>\n"
                        f"⚠️ Требует проверки: {total_checks_needed}\n"
                        f"🟢 Чистых: {total_clean}\n\n"
                        f"💾 Файл сохранён"
                    )
                
            except Exception as e_final:
                logger_func(f"ОШИБКА СОХРАНЕНИЯ: {e_final}")
                logger_func(traceback.format_exc())
        else:
            logger_func(f"\n{'='*60}")
            logger_func("НЕТ ДАННЫХ ДЛЯ СОХРАНЕНИЯ")
            logger_func(f"  Всего проверено: {len(clients_data)}")
            logger_func(f"  Все клиенты чистые или дела слишком старые: {total_clean}")
            logger_func(f"{'='*60}")
            
            # 📱 Уведомление о завершении (все чистые)
            if TELEGRAM_AVAILABLE and is_bot_available():
                send_notification_sync(
                    f"✅ <b>БАНКРОТСТВО ЗАВЕРШЕНО</b>\n\n"
                    f"📝 Всего проверено: {len(clients_data)}\n"
                    f"🟢 Все клиенты чистые!\n\n"
                    f"Нет данных для сохранения"
                )

        return output_file

    except WebDriverException as e:
        logger_func(f"Ошибка вебдрайвера: {e}")
        logger_func(traceback.format_exc())
        
        # 📱 Критическое уведомление
        if TELEGRAM_AVAILABLE and is_bot_available():
            send_notification_sync(
                f"🚨 <b>БАНКРОТСТВО УПАЛО</b>\n\n"
                f"❌ Ошибка браузера: {str(e)[:100]}\n\n"
                f"Проверьте ChromeDriver"
            )
        return None
    except Exception as e:
        logger_func(f"Глобальная ошибка: {e}")
        logger_func(traceback.format_exc())
        
        # 📱 Критическое уведомление
        if TELEGRAM_AVAILABLE and is_bot_available():
            send_notification_sync(
                f"🚨 <b>БАНКРОТСТВО УПАЛО</b>\n\n"
                f"❌ Критическая ошибка: {str(e)[:100]}\n\n"
                f"Проверьте логи программы"
            )
        return None
    finally:
        if driver:
            logger_func("Закрытие браузера через 3 секунды...")
            time.sleep(3)
            driver.quit()
            logger_func("Браузер закрыт.")